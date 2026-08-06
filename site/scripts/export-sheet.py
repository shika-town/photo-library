# -*- coding: utf-8 -*-
"""現在のサイト内容を、運用シート（Excel）に書き出します。

  使い方:  python3 scripts/export-sheet.py     ← site フォルダ直下で実行

書き出したファイルを Google スプレッドシートにアップロードすれば、
そのまま「今のサイトの内容」から編集を始められます。
"""
import json, io, os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT  = os.path.join(BASE, '..', '運用シート_SHIKA_PHOTO_LIBRARY.xlsx')

top   = json.load(io.open(os.path.join(BASE, 'data', 'photos.json'), encoding='utf-8'))
spots = json.load(io.open(os.path.join(BASE, 'data', 'spots.json'),  encoding='utf-8'))
lib   = json.load(io.open(os.path.join(BASE, 'data', 'library.json'),encoding='utf-8'))

# サムネイルのパスから写真idを引く
photo_by_thumb = {p['thumb']: p['id'] for p in lib['photos']}

FONT = 'Yu Gothic'
NAVY, NAVY2, SOFT, GREY, LINE = '0B4F82', '08395E', 'E7F0F7', 'F6F8FA', 'D8E0E8'
f_base = Font(name=FONT, size=10)
f_bold = Font(name=FONT, size=10, bold=True)
f_head = Font(name=FONT, size=10, bold=True, color='FFFFFF')
f_in   = Font(name=FONT, size=10, color='0000FF')   # 入力してよい欄
f_lock = Font(name=FONT, size=10, color='808080')   # 触らない欄
f_title= Font(name=FONT, size=14, bold=True, color=NAVY)
f_sub  = Font(name=FONT, size=10, color='5A6C7D')
f_note = Font(name=FONT, size=9,  color='8A6D3B')
fill_head  = PatternFill('solid', fgColor=NAVY)
fill_head2 = PatternFill('solid', fgColor=NAVY2)
fill_grey  = PatternFill('solid', fgColor=GREY)
fill_warn  = PatternFill('solid', fgColor='FFF7D6')
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def sheet(name, headers, widths, locked=(), rows=(), extra=400):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = f_head
        c.fill = fill_head2 if h in locked else fill_head
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'B2'
    for r, row in enumerate(rows, 2):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=i, value=row.get(h, ''))
            c.font = f_lock if h in locked else f_in
            c.border = border
            c.alignment = Alignment(vertical='center', wrap_text=False)
            if h in locked:
                c.fill = fill_grey
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(headers)), len(rows) + 1)
    return ws


# =====================================================================
# 1. はじめに
# =====================================================================
ws = wb.active
ws.title = 'はじめに'
ws.sheet_view.showGridLines = False
for col, w in (('A', 3), ('B', 22), ('C', 16), ('D', 78)):
    ws.column_dimensions[col].width = w

def put(cell, v, font=f_base, fill=None, wrap=False):
    c = ws[cell]; c.value = v; c.font = font
    if fill: c.fill = fill
    c.alignment = Alignment(vertical='center', wrap_text=wrap)

put('B2', 'SHIKA PHOTO LIBRARY ／ 運用シート', f_title)
put('B3', 'このシートを編集すると、サイトの内容が変わります。', f_sub)
put('B4', '書き出し日時: ' + datetime.datetime.now().strftime('%Y-%m-%d %H:%M'), f_sub)

put('B6', '■ シートの役割', f_bold)
roles = [
    ('写真',       'サイトに載せる写真の一覧です。1行＝写真1枚。追加・削除はこのシートで行います。'),
    ('スポット',   '観光地ごとのページの内容です。1行＝1ページ。説明文もここで直します。'),
    ('サイト設定', 'トップページのキャッチコピーやお問い合わせ先など、サイト共通の設定です。'),
    ('シーン',     'トップページの「おすすめシーン」6枚の設定です。'),
    ('季節',       'トップページの「四季折々の志賀町」4枚の設定です。'),
    ('選択肢',     '各シートのドロップダウンの元になるリストです。増やしたいときはここに追記します。'),
]
r = 7
for a, b in roles:
    put('B%d' % r, a, f_bold); put('D%d' % r, b)
    ws.merge_cells('D%d:D%d' % (r, r))
    r += 1

put('B15', '■ 色のきまり', f_bold)
put('B16', '青い文字', f_bold); put('C16', '編集OK', f_in);   put('D16', '自由に書き換えてかまいません。')
put('B17', '灰色の文字', f_bold); put('C17', '触らない', f_lock); put('D17', 'システムが使う欄です。編集するとサイトが正しく作られません。')
put('B18', '黄色の背景', f_bold); put('C18', '要確認', f_base, fill_warn); put('D18', '公開前に必ず確認してほしい欄です。')
for x in ('C16', 'C17', 'C18'):
    ws[x].alignment = Alignment(horizontal='center', vertical='center'); ws[x].border = border

put('B20', '■ 更新の流れ', f_bold)
flow = [
    ('1', 'このシートを直します。写真を増やすときは「写真」シートに行を追加します。'),
    ('2', '写真の実物は Google ドライブに入れ、共有リンクのIDを driveFileId 欄に貼ります。'),
    ('3', '公開したい行は publish 欄を TRUE にします。FALSE の行はサイトに出ません。'),
    ('4', '保存すると、あとは自動でサイトに反映されます（数分かかります）。'),
    ('5', 'サイトを開いて、意図どおりか確認します。'),
]
r = 21
for no, txt in flow:
    put('B%d' % r, no, f_bold); ws['B%d' % r].alignment = Alignment(horizontal='center')
    ws.merge_cells('C%d:D%d' % (r, r)); put('C%d' % r, txt)
    r += 1

put('B27', '■ 注意', f_bold)
notes = [
    '・列の順番を入れ替えたり、列を削除したりしないでください。サイトが作られなくなります。',
    '・行の追加・削除、並べ替えは自由です。並び順はそのままサイトの表示順になります。',
    '・id 欄は写真やページを見分けるための番号です。一度決めたら変えないでください。',
    '・複数人で同時に編集できます。誰がいつ直したかは updatedBy / updatedAt 欄に残してください。',
    '・削除した写真は元に戻せません。迷ったら publish を FALSE にして非表示にしてください。',
]
r = 28
for n in notes:
    ws.merge_cells('B%d:D%d' % (r, r)); put('B%d' % r, n)
    r += 1

# =====================================================================
# 2. 写真
# =====================================================================
PH = ['id','title','spot','area','season','tags','roles','sortOrder','driveFileId','localPath',
      'focus','photographer','copyright','description','publish','updatedBy','updatedAt']
PW = [9, 30, 26, 13, 9, 26, 20, 11, 34, 34, 12, 13, 12, 40, 10, 13, 13]
# トップの新着写真は library.json の先頭に同じ順で並んでいる（build-library.py の仕様）
N_TOP = len(top['photos'])

# 初回の書き出しに限り、トップとスポットで同じ元写真を使っている重複を整理する。
#   ・スポット側の行を落とし、トップ側の行にスポット名を持たせる
#   ・こうしないと、同じ写真がギャラリーに2回並ぶ
# 落とすスポット側の行 → 代わりに使うトップ側の行
DEDUP_DROP = {
    ('hatago',   '岩の間に沈む夕陽'):     'P0001',
    ('gate',     '上空から見た入り江'):    'P0002',
    ('sakuragai','手のひらのさくら貝'):    'P0004',
    ('yase',     '日本海に切り立つ断崖'):  'P0005',
    ('fukura',   '日和山に立つ白亜の灯台'): 'P0006',
    ('bench',    '冬のイルミネーション'):  'P0007',
}
RESPOT = {   # トップ側の行に、正しいスポット名を割り当て直す
    'P0004': '道の駅とぎ海街道・さくら貝資料館',
    'P0005': 'ヤセの断崖・義経の舟隠し',
    'P0006': '旧福浦灯台',
    'P0007': '世界一長いベンチ・増穂浦海岸',
}
area_by_spot = {s['name']: s['area'] for s in spots['spots']}

rows = []
for i, p in enumerate(lib['photos']):
    if p['spotId'] and (p['spotId'], p['title']) in DEDUP_DROP:
        continue
    roles = ['新着'] if i < N_TOP else []
    spot = RESPOT.get(p['id'], p['spot'])
    rows.append({
        'id': p['id'], 'title': p['title'], 'spot': spot,
        'area': area_by_spot.get(spot, p['area']),
        'season': p.get('season', ''), 'tags': ','.join(p['tags']),
        'roles': ','.join(roles), 'sortOrder': '', 'driveFileId': '', 'localPath': '_source/photos/%s.jpg' % p['id'],
        'focus': '中央', 'photographer': p['photographer'], 'copyright': p['credit'],
        'description': p.get('alt', ''), 'publish': 'TRUE',
        'updatedBy': '', 'updatedAt': '',
    })

# シーン・季節に使っている写真は library.json に入っていないため、行として追加する
n = len(lib['photos'])
scene_row, season_row = {}, {}
for sc in top['scenes']:
    n += 1; pid = 'P%04d' % n
    scene_row[sc['label']] = pid
    rows.append({'id': pid, 'title': sc['alt'], 'spot': '', 'area': '', 'season': '',
                 'tags': sc['query'], 'roles': 'シーン', 'sortOrder': 10 * len(scene_row), 'driveFileId': '',
                 'localPath': '_source/photos/scene-%02d.jpg' % len(scene_row), 'focus': '中央', 'photographer': '志賀町',
                 'copyright': '© 志賀町', 'description': sc['alt'], 'publish': 'TRUE',
                 'updatedBy': '', 'updatedAt': ''})
for se in top['seasons']:
    n += 1; pid = 'P%04d' % n
    season_row[se['ja']] = pid
    rows.append({'id': pid, 'title': se['alt'], 'spot': '', 'area': '', 'season': se['ja'],
                 'tags': se['query'], 'roles': '季節', 'sortOrder': 10 * len(season_row), 'driveFileId': '',
                 'localPath': '_source/photos/season-%02d.jpg' % len(season_row), 'focus': '中央', 'photographer': '志賀町',
                 'copyright': '© 志賀町', 'description': se['alt'], 'publish': 'TRUE',
                 'updatedBy': '', 'updatedAt': ''})
# 元のギャラリー順を保つように sortOrder を振り、落とした行は代わりの行に読み替える
by_id_row = {r['id']: r for r in rows}
hero_of = {}
for sp in spots['spots']:
    order = []
    for ph in sp['photos']:
        pid = photo_by_thumb.get(ph['thumb'].replace('../', ''), '')
        pid = DEDUP_DROP.get((sp['id'], ph['caption']), pid)
        if pid and pid in by_id_row and pid not in order:
            order.append(pid)
    for i, pid in enumerate(order, 1):
        by_id_row[pid]['sortOrder'] = i * 10
    # ヒーローは spots.json の heroCaption で指定された写真を使う
    want = sp.get('heroCaption')
    hid = ''
    if want:
        hid = DEDUP_DROP.get((sp['id'], want), '')
        if not hid:
            for ph in sp['photos']:
                if ph['caption'] == want:
                    hid = photo_by_thumb.get(ph['thumb'].replace('../', ''), '')
                    break
    hero_of[sp['id']] = hid or (order[0] if order else '')
n_rest = 0
for r in rows:
    if r['sortOrder'] == '':
        n_rest += 1
        r['sortOrder'] = n_rest * 10

ws = sheet('写真', PH, PW, locked=('id',), rows=rows)
ws['A1'].comment = Comment('写真を見分ける番号です。新しい行は P0066 のように続けて付けてください。', 'SHIKA PHOTO LIBRARY', height=90, width=280)
ws['I1'].comment = Comment('Google ドライブの共有リンク\nhttps://drive.google.com/file/d/【ここ】/view\nの【ここ】を貼り付けます。', 'SHIKA PHOTO LIBRARY', height=110, width=300)
ws['J1'].comment = Comment('ドライブを使わず、リポジトリ内の画像を指定する場合のパスです。\ndriveFileId が空のときだけ使われます。', 'SHIKA PHOTO LIBRARY', height=100, width=320)
ws['H1'].comment = Comment('小さいほど先に表示されます。並べ替えたいときは数字を変えてください。', 'SHIKA PHOTO LIBRARY', height=90, width=320)
ws['G1'].comment = Comment('トップページのどこに出すか。新着／シーン／季節 をカンマ区切りで。\n空欄ならスポットページにだけ出ます。', 'SHIKA PHOTO LIBRARY', height=100, width=320)
ws['K1'].comment = Comment('切り抜きの基準位置。中央／上寄り／下寄り から選びます。\n人物の顔が切れるときに調整してください。', 'SHIKA PHOTO LIBRARY', height=100, width=320)
for r in range(2, len(rows) + 2):
    ws.cell(row=r, column=PH.index('publish') + 1).fill = fill_warn   # publish

# =====================================================================
# 3. スポット
# =====================================================================
SP = ['id','name','kana','area','address','catch','body1','body2','body3','body4',
      'tags','heroPhotoId','info1_label','info1_value','info2_label','info2_value',
      'info3_label','info3_value','info4_label','info4_value','info5_label','info5_value',
      'info6_label','info6_value',
      'source1_label','source1_url','source2_label','source2_url','publish','updatedBy','updatedAt']
SW = [11, 28, 30, 13, 30, 34, 52, 52, 52, 52, 26, 12,
      16, 34, 16, 34, 16, 34, 16, 34, 16, 34, 16, 34, 18, 40, 18, 40, 10, 13, 13]
srows = []
photo_by_thumb = {p['thumb']: p['id'] for p in lib['photos']}
for s in spots['spots']:
    row = {'id': s['id'], 'name': s['name'], 'kana': s['kana'], 'area': s['area'],
           'address': s['address'], 'catch': s['catch'], 'tags': ','.join(s['tags']),
           'heroPhotoId': hero_of.get(s['id'], ''),
           'publish': 'TRUE', 'updatedBy': '', 'updatedAt': ''}
    for i, b in enumerate(s['body'][:4], 1):
        row['body%d' % i] = b
    for i, inf in enumerate(s['info'][:6], 1):
        row['info%d_label' % i] = inf['label']; row['info%d_value' % i] = inf['value']
    for i, sc in enumerate(s['sources'][:2], 1):
        row['source%d_label' % i] = sc['label']; row['source%d_url' % i] = sc['url']
    srows.append(row)
ws = sheet('スポット', SP, SW, locked=('id',), rows=srows)
ws['L1'].comment = Comment('ページ上部の大きな写真に使う写真のidです。\n「写真」シートのidを入れてください。', 'SHIKA PHOTO LIBRARY', height=100, width=300)
ws['G1'].comment = Comment('説明文の段落です。body1から順に表示されます。\n使わない段落は空欄のままでかまいません。', 'SHIKA PHOTO LIBRARY', height=100, width=320)
for r in range(2, len(srows) + 2):
    ws.cell(row=r, column=SP.index('publish') + 1).fill = fill_warn

# =====================================================================
# 4. サイト設定
# =====================================================================
ws = wb.create_sheet('サイト設定')
for i, h in enumerate(['key', '項目', '値', '説明'], 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = f_head; c.fill = fill_head if i != 1 else fill_head2
    c.border = border; c.alignment = Alignment(horizontal='center', vertical='center')
for col, w in (('A', 20), ('B', 26), ('C', 52), ('D', 54)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'C2'

SITE_ROWS = [
    ('siteName',    'サイト名（英語）',       top['site']['name'],   'ヘッダーのロゴに出ます。'),
    ('siteNameJa',  'サイト名（日本語）',     top['site']['nameJa'], 'ロゴの下の小さな文字です。'),
    ('tagline',     '説明文',                 top['site']['tagline'],'検索結果やSNSに出る説明です。'),
    ('heroTitle1',  'キャッチコピー1行目',    top['hero']['title'][0], 'トップの大きな文字（1行目）。'),
    ('heroTitle2',  'キャッチコピー2行目',    top['hero']['title'][1] if len(top['hero']['title']) > 1 else '', 'トップの大きな文字（2行目）。'),
    ('heroPhotoId', 'トップの大きな写真',     next((r['id'] for r in rows if r['localPath'].endswith('bench-01-lg.jpg')), 'P0009'), '「写真」シートのidを入れます。'),
    ('heroCaption', '写真の撮影地表示',       top['hero']['caption'], '大きな写真の右下に出ます。'),
    ('heroTags',    '人気のタグ',             ','.join(top['heroTags']), '検索ボックスの下に並ぶタグ。カンマ区切り。'),
    ('copyright',   '著作権表示',             top['site']['copyright'], 'フッターの一番下に出ます。'),
    ('kankouUrl',   '観光情報サイト',         'https://shika-guide.jp', 'フッターのボタンのリンク先。'),
    ('instagramUrl','Instagram',              'https://www.instagram.com/shikatown_official/', 'フッターのボタンのリンク先。'),
    ('contactMail', 'お問い合わせ先メール',   'shokan@town.shika.lg.jp', '規約ページなどに表示されます。'),
    ('contactDept', '担当部署名',             '志賀町役場 商工観光課', '規約ページなどに表示されます。'),
    ('formUrl',     'お問い合わせフォーム',   '', 'Googleフォームの埋め込みURL。空欄なら準備中の表示になります。'),
]
for r, (k, label, v, desc) in enumerate(SITE_ROWS, 2):
    ws.cell(row=r, column=1, value=k).font = f_lock
    ws.cell(row=r, column=1).fill = fill_grey
    ws.cell(row=r, column=2, value=label).font = f_bold
    ws.cell(row=r, column=3, value=v).font = f_in
    ws.cell(row=r, column=4, value=desc).font = f_base
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = border
        ws.cell(row=r, column=i).alignment = Alignment(vertical='center')

# =====================================================================
# 5. シーン / 6. 季節
# =====================================================================
sheet('シーン', ['label','query','photoId','sortOrder','publish'], [16, 18, 12, 12, 10],
      locked=(), rows=[{'label': s['label'], 'query': s['query'],
                        'photoId': scene_row.get(s['label'], ''), 'sortOrder': i + 1, 'publish': 'TRUE'}
                       for i, s in enumerate(top['scenes'])])
sheet('季節', ['ja','en','query','photoId','sortOrder','publish'], [10, 14, 14, 12, 12, 10],
      locked=(), rows=[{'ja': s['ja'], 'en': s['en'], 'query': s['query'],
                        'photoId': season_row.get(s['ja'], ''), 'sortOrder': i + 1, 'publish': 'TRUE'}
                       for i, s in enumerate(top['seasons'])])

# =====================================================================
# 6.5 フォーム回答（Googleフォームを連携すると自動で埋まります）
# =====================================================================
FR = ['タイムスタンプ','メールアドレス','写真','タイトル','スポット','季節','タグ','撮影者','備考',
      '承認','id','sortOrder','focus']
FW = [18, 26, 40, 28, 26, 9, 24, 14, 34, 10, 10, 11, 12]
fs = sheet('フォーム回答', FR, FW, locked=('タイムスタンプ','メールアドレス','写真','id'), rows=[
  {'タイムスタンプ': '2026/08/10 9:15:22', 'メールアドレス': 'shokan@town.shika.lg.jp',
   '写真': 'https://drive.google.com/open?id=（フォームが自動で入れます）',
   'タイトル': '（記入例）夕暮れの巌門', 'スポット': '巌門', '季節': '秋',
   'タグ': '夕陽,海,岩', '撮影者': '志賀町', '備考': '遊歩道から撮影',
   '承認': 'FALSE', 'id': '', 'sortOrder': '', 'focus': '中央'},
])
fs['A1'].comment = Comment('Googleフォームと連携すると、この行より下に回答が自動で増えていきます。\n'
                           'グレーの列は触らないでください。', 'SHIKA PHOTO LIBRARY', height=110, width=320)
fs['J1'].comment = Comment('内容を確認して、サイトに載せてよければ TRUE にします。\n'
                           'FALSE のあいだはサイトに出ません。', 'SHIKA PHOTO LIBRARY', height=100, width=320)
fs['M1'].comment = Comment('切り抜きの基準位置。人物の顔が切れるときに\n「上寄り」などに変えてください。',
                           'SHIKA PHOTO LIBRARY', height=90, width=320)
for r in range(2, 3):
    fs.cell(row=r, column=FR.index('承認') + 1).fill = fill_warn
fs.cell(row=4, column=1, value='※ 1行目の記入例は、フォームを連携したら削除してください。').font = f_note

# =====================================================================
# 7. 選択肢
# =====================================================================
ms = wb.create_sheet('選択肢')
ms.sheet_view.showGridLines = False
lists = [
    ('エリア',  sorted({s['area'] for s in spots['spots']}), 16),
    ('季節',    ['春', '夏', '秋', '冬', '通年'], 12),
    ('公開',    ['TRUE', 'FALSE'], 12),
    ('切り抜き', ['中央', '上寄り', '下寄り'], 14),
    ('役割',    ['新着', 'シーン', '季節'], 14),
    ('スポット', [s['name'] for s in spots['spots']], 30),
]
col = 1
for title, values, w in lists:
    c = ms.cell(row=1, column=col, value=title)
    c.font = f_head; c.fill = fill_head; c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center')
    ms.column_dimensions[get_column_letter(col)].width = w
    for i, v in enumerate(values):
        cc = ms.cell(row=2 + i, column=col, value=v); cc.font = f_in; cc.border = border
    col += 2
ms.cell(row=len(lists[-1][1]) + 4, column=1,
        value='※ 選択肢を増やすときは、各リストの最終行の下に追記してください。').font = f_note

for s in wb.worksheets:
    s.sheet_properties.tabColor = NAVY

wb.save(OUT)
print('書き出しました:', os.path.normpath(OUT))
print('  写真 %d行 / スポット %d行 / シーン %d行 / 季節 %d行'
      % (len(rows), len(srows), len(top['scenes']), len(top['seasons'])))
